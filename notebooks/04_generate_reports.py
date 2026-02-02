import marimo

__generated_with = "0.9.0"
app = marimo.App(width="medium")


@app.cell
def __():
    import marimo as mo
    return mo,


@app.cell
def __(mo):
    mo.md(
        r"""
        # 04 - Generate Reports

        Export normalized data to formatted Excel reports.

        ## Features
        - Multiple sheets per workbook
        - Formatted headers and styling
        - Summary statistics
        - Date-stamped output files
        """
    )
    return


@app.cell
def __():
    import duckdb
    from pathlib import Path
    from datetime import datetime
    import pandas as pd
    return Path, datetime, duckdb, pd


@app.cell
def __(Path, duckdb):
    # Connect to database
    PROJECT_ROOT = Path(__file__).parent.parent
    DB_PATH = PROJECT_ROOT / "data" / "processed" / "analytics.duckdb"
    OUTPUT_DIR = PROJECT_ROOT / "data" / "output"
    TEMPLATE_DIR = PROJECT_ROOT / "templates"

    # Ensure output directory exists
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    conn = duckdb.connect(str(DB_PATH))
    return DB_PATH, OUTPUT_DIR, PROJECT_ROOT, TEMPLATE_DIR, conn


@app.cell
def __(mo):
    mo.md("## Report Configuration")
    return


@app.cell
def __(mo):
    # Report type selection
    report_type = mo.ui.dropdown(
        options=["Monthly Summary", "Customer Report", "Product Performance", "Full Export"],
        value="Monthly Summary",
        label="Select report type:"
    )
    report_type
    return report_type,


@app.cell
def __(datetime, mo):
    # Date for filename
    today = datetime.now().strftime("%Y%m%d")
    mo.md(f"Report date: **{today}**")
    return today,


@app.cell
def __(mo):
    mo.md("## Generate Report")
    return


@app.cell
def __(mo):
    generate_button = mo.ui.run_button(label="Generate Report")
    generate_button
    return generate_button,


@app.cell
def __(
    OUTPUT_DIR,
    conn,
    datetime,
    generate_button,
    mo,
    pd,
    report_type,
    today,
):
    mo.stop(not generate_button.value)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        mo.md("**Error:** openpyxl is required. Install with: `pip install openpyxl`")
        raise

    def style_header(ws):
        """Apply header styling to the first row."""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    def auto_width(ws):
        """Auto-adjust column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def df_to_sheet(wb, df, sheet_name):
        """Add a DataFrame to a workbook as a new sheet."""
        ws = wb.create_sheet(title=sheet_name)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        style_header(ws)
        auto_width(ws)
        return ws

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    report_name = report_type.value.lower().replace(" ", "_")
    output_file = OUTPUT_DIR / f"{report_name}_{today}.xlsx"

    if report_type.value == "Monthly Summary":
        df = conn.execute("SELECT * FROM vw_monthly_summary").fetchdf()
        df_to_sheet(wb, df, "Monthly Summary")

    elif report_type.value == "Customer Report":
        df = conn.execute("SELECT * FROM vw_customer_report").fetchdf()
        df_to_sheet(wb, df, "Customer Report")

    elif report_type.value == "Product Performance":
        df = conn.execute("SELECT * FROM vw_product_performance").fetchdf()
        df_to_sheet(wb, df, "Product Performance")

    elif report_type.value == "Full Export":
        # Export all views and dimension tables
        views = [
            ("Monthly Summary", "SELECT * FROM vw_monthly_summary"),
            ("Customer Report", "SELECT * FROM vw_customer_report"),
            ("Product Performance", "SELECT * FROM vw_product_performance"),
            ("Customers", "SELECT * FROM dim_customers"),
            ("Products", "SELECT * FROM dim_products"),
            ("Orders", "SELECT * FROM fact_orders"),
        ]
        for sheet_name, query in views:
            df = conn.execute(query).fetchdf()
            df_to_sheet(wb, df, sheet_name)

    # Save workbook
    wb.save(output_file)

    mo.md(f"""
    **Report generated successfully!**

    - File: `{output_file.name}`
    - Location: `data/output/`
    - Sheets: {len(wb.sheetnames)}
    """)
    return (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
        Workbook,
        adjusted_width,
        auto_width,
        cell,
        column,
        column_letter,
        dataframe_to_rows,
        df,
        df_to_sheet,
        max_length,
        output_file,
        query,
        report_name,
        sheet_name,
        style_header,
        views,
        wb,
        ws,
    )


@app.cell
def __(mo):
    mo.md("## Preview Generated Report")
    return


@app.cell
def __(generate_button, mo, output_file, pd):
    mo.stop(not generate_button.value)

    # Show preview of generated file
    preview_df = pd.read_excel(output_file, sheet_name=None)

    tabs_content = {}
    for sheet_name, df in preview_df.items():
        tabs_content[sheet_name] = mo.ui.table(df.head(20))

    mo.ui.tabs(tabs_content)
    return df, preview_df, sheet_name, tabs_content


@app.cell
def __(mo):
    mo.md("## Batch Export")
    return


@app.cell
def __(mo):
    mo.md(
        """
        Generate all reports at once for monthly distribution.
        """
    )
    return


@app.cell
def __(mo):
    batch_button = mo.ui.run_button(label="Generate All Reports")
    batch_button
    return batch_button,


@app.cell
def __(
    OUTPUT_DIR,
    auto_width,
    batch_button,
    conn,
    dataframe_to_rows,
    df_to_sheet,
    mo,
    style_header,
    today,
):
    from openpyxl import Workbook as WB

    mo.stop(not batch_button.value)

    reports_generated = []

    # Report 1: Monthly Summary
    wb1 = WB()
    wb1.remove(wb1.active)
    df1 = conn.execute("SELECT * FROM vw_monthly_summary").fetchdf()
    df_to_sheet(wb1, df1, "Monthly Summary")
    file1 = OUTPUT_DIR / f"monthly_summary_{today}.xlsx"
    wb1.save(file1)
    reports_generated.append(file1.name)

    # Report 2: Customer Report
    wb2 = WB()
    wb2.remove(wb2.active)
    df2 = conn.execute("SELECT * FROM vw_customer_report").fetchdf()
    df_to_sheet(wb2, df2, "Customer Report")
    file2 = OUTPUT_DIR / f"customer_report_{today}.xlsx"
    wb2.save(file2)
    reports_generated.append(file2.name)

    # Report 3: Product Performance
    wb3 = WB()
    wb3.remove(wb3.active)
    df3 = conn.execute("SELECT * FROM vw_product_performance").fetchdf()
    df_to_sheet(wb3, df3, "Product Performance")
    file3 = OUTPUT_DIR / f"product_performance_{today}.xlsx"
    wb3.save(file3)
    reports_generated.append(file3.name)

    mo.md(f"""
    **Batch export complete!**

    Generated {len(reports_generated)} reports:
    """ + "\n".join([f"- `{r}`" for r in reports_generated]))
    return (
        WB,
        df1,
        df2,
        df3,
        file1,
        file2,
        file3,
        reports_generated,
        wb1,
        wb2,
        wb3,
    )


@app.cell
def __(mo):
    mo.md(
        r"""
        ## Workflow Complete

        You have successfully:
        1. Ingested Excel data into DuckDB
        2. Explored and validated the data
        3. Created a normalized schema
        4. Generated formatted Excel reports

        All data processing happened locally with no network access required.
        """
    )
    return


if __name__ == "__main__":
    app.run()
