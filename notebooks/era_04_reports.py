import marimo

__generated_with = "0.9.0"
app = marimo.App(width="medium")


@app.cell
def __():
    import marimo as mo
    return mo,


@app.cell
def __(mo):
    mo.md(
        r"""
        # ERA 04 - Generate Reports

        Export ERA summary tables to formatted Excel reports.

        ## Standard ERA Tables
        - Table 1: Sample Summary
        - Table 2: Detection Summary
        - Table 3: Screening Level Comparison
        - Table 4: COPC Summary
        - Table 5: Exposure Point Concentrations
        """
    )
    return


@app.cell
def __():
    import duckdb
    from pathlib import Path
    from datetime import datetime
    import pandas as pd
    import numpy as np
    from scipy import stats
    return Path, datetime, duckdb, np, pd, stats


@app.cell
def __(Path, duckdb):
    PROJECT_ROOT = Path(__file__).parent.parent
    DB_PATH = PROJECT_ROOT / "data" / "processed" / "analytics.duckdb"
    OUTPUT_DIR = PROJECT_ROOT / "data" / "output"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    conn = duckdb.connect(str(DB_PATH))
    return DB_PATH, OUTPUT_DIR, PROJECT_ROOT, conn


@app.cell
def __(mo):
    mo.md("## Report Configuration")
    return


@app.cell
def __(mo):
    site_name = mo.ui.text(
        value="Former Manufacturing Facility",
        label="Site Name:"
    )
    site_name
    return site_name,


@app.cell
def __(mo):
    report_date = mo.ui.text(
        value=datetime.now().strftime("%B %Y"),
        label="Report Date:"
    )
    report_date
    return report_date,


@app.cell
def __(datetime, mo):
    from datetime import datetime as dt
    today = dt.now().strftime("%Y%m%d")
    mo.md(f"Output file date stamp: `{today}`")
    return dt, today


@app.cell
def __(mo):
    generate_btn = mo.ui.run_button(label="Generate ERA Report")
    generate_btn
    return generate_btn,


@app.cell
def __(
    OUTPUT_DIR,
    conn,
    generate_btn,
    mo,
    np,
    pd,
    report_date,
    site_name,
    stats,
    today,
):
    mo.stop(not generate_btn.value)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    exceed_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    def auto_width(ws):
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

    def add_title(ws, title, site, date):
        ws["A1"] = title
        ws["A1"].font = title_font
        ws["A2"] = f"Site: {site}"
        ws["A3"] = f"Date: {date}"
        return 5  # Starting row for data

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)

    # ========================================
    # Table 1: Sample Summary
    # ========================================
    ws1 = wb.create_sheet("Table 1 - Samples")
    start_row = add_title(ws1, "Table 1: Sample Summary", site_name.value, report_date.value)

    sample_query = """
    SELECT
        s.sample_id as "Sample ID",
        s.location_id as "Location",
        l.location_type as "Type",
        s.sample_date as "Date",
        m.matrix_name as "Matrix",
        s.depth_top_ft as "Top (ft)",
        s.depth_bottom_ft as "Bottom (ft)",
        COUNT(r.result_id) as "Analytes"
    FROM fact_samples s
    LEFT JOIN dim_locations l ON s.location_id = l.location_id
    LEFT JOIN dim_matrix m ON s.matrix_code = m.matrix_code
    LEFT JOIN fact_results r ON s.sample_id = r.sample_id
    GROUP BY s.sample_id, s.location_id, l.location_type, s.sample_date, m.matrix_name, s.depth_top_ft, s.depth_bottom_ft
    ORDER BY s.sample_date, s.location_id
    """
    sample_df = conn.execute(sample_query).fetchdf()

    for r_idx, row in enumerate(dataframe_to_rows(sample_df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border

    style_header(ws1, start_row)
    auto_width(ws1)

    # ========================================
    # Table 2: Detection Summary
    # ========================================
    ws2 = wb.create_sheet("Table 2 - Detection")
    start_row = add_title(ws2, "Table 2: Detection Summary by Analyte", site_name.value, report_date.value)

    detection_query = """
    SELECT
        a.analyte_name as "Analyte",
        a.analyte_group as "Group",
        m.matrix_name as "Matrix",
        COUNT(*) as "N Samples",
        SUM(CASE WHEN r.detect_flag = 'Y' THEN 1 ELSE 0 END) as "N Detect",
        ROUND(100.0 * SUM(CASE WHEN r.detect_flag = 'Y' THEN 1 ELSE 0 END) / COUNT(*), 1) as "Detect %",
        MIN(CASE WHEN r.detect_flag = 'Y' THEN r.result_value END) as "Min",
        MAX(CASE WHEN r.detect_flag = 'Y' THEN r.result_value END) as "Max",
        r.result_unit as "Unit"
    FROM fact_results r
    JOIN fact_samples s ON r.sample_id = s.sample_id
    LEFT JOIN dim_analytes a ON r.cas_rn = a.cas_rn
    LEFT JOIN dim_matrix m ON s.matrix_code = m.matrix_code
    GROUP BY a.analyte_name, a.analyte_group, m.matrix_name, r.result_unit
    HAVING SUM(CASE WHEN r.detect_flag = 'Y' THEN 1 ELSE 0 END) > 0
    ORDER BY a.analyte_group, a.analyte_name
    """
    detection_df = conn.execute(detection_query).fetchdf()

    for r_idx, row in enumerate(dataframe_to_rows(detection_df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border

    style_header(ws2, start_row)
    auto_width(ws2)

    # ========================================
    # Table 3: Screening Comparison
    # ========================================
    ws3 = wb.create_sheet("Table 3 - Screening")
    start_row = add_title(ws3, "Table 3: Screening Level Comparison", site_name.value, report_date.value)

    screening_query = """
    SELECT
        a.analyte_name as "Analyte",
        m.matrix_name as "Matrix",
        MAX(r.result_value) as "Max Conc",
        r.result_unit as "Unit",
        CASE s.matrix_code
            WHEN 'SO' THEN sl.rsl_residential_soil_mg_kg
            WHEN 'GW' THEN sl.rsl_residential_tap_ug_l
        END as "RSL",
        ROUND(MAX(r.result_value) / NULLIF(
            CASE s.matrix_code
                WHEN 'SO' THEN sl.rsl_residential_soil_mg_kg
                WHEN 'GW' THEN sl.rsl_residential_tap_ug_l
            END, 0), 2) as "HQ",
        CASE
            WHEN MAX(r.result_value) > CASE s.matrix_code
                WHEN 'SO' THEN sl.rsl_residential_soil_mg_kg
                WHEN 'GW' THEN sl.rsl_residential_tap_ug_l
            END THEN 'EXCEEDS'
            ELSE 'Below'
        END as "Status",
        sl.carcinogen as "Carcinogen"
    FROM fact_results r
    JOIN fact_samples s ON r.sample_id = s.sample_id
    LEFT JOIN dim_analytes a ON r.cas_rn = a.cas_rn
    LEFT JOIN dim_matrix m ON s.matrix_code = m.matrix_code
    LEFT JOIN ref_screening_levels sl ON r.cas_rn = sl.cas_rn
    WHERE r.detect_flag = 'Y'
    AND sl.cas_rn IS NOT NULL
    GROUP BY a.analyte_name, m.matrix_name, s.matrix_code, r.result_unit,
             sl.rsl_residential_soil_mg_kg, sl.rsl_residential_tap_ug_l, sl.carcinogen
    ORDER BY
        CASE WHEN MAX(r.result_value) > CASE s.matrix_code
            WHEN 'SO' THEN sl.rsl_residential_soil_mg_kg
            WHEN 'GW' THEN sl.rsl_residential_tap_ug_l
        END THEN 0 ELSE 1 END,
        MAX(r.result_value) / NULLIF(CASE s.matrix_code
            WHEN 'SO' THEN sl.rsl_residential_soil_mg_kg
            WHEN 'GW' THEN sl.rsl_residential_tap_ug_l
        END, 0) DESC
    """
    screening_df = conn.execute(screening_query).fetchdf()

    for r_idx, row in enumerate(dataframe_to_rows(screening_df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            # Highlight exceedances
            if c_idx == 7 and value == 'EXCEEDS':
                for c in range(1, 9):
                    ws3.cell(row=r_idx, column=c).fill = exceed_fill

    style_header(ws3, start_row)
    auto_width(ws3)

    # ========================================
    # Table 4: COPC Summary
    # ========================================
    ws4 = wb.create_sheet("Table 4 - COPCs")
    start_row = add_title(ws4, "Table 4: Chemicals of Potential Concern (COPCs)", site_name.value, report_date.value)

    copc_query = """
    SELECT
        a.analyte_name as "Analyte",
        a.analyte_group as "Group",
        m.matrix_name as "Matrix",
        COUNT(*) as "N Exceed",
        MAX(r.result_value) as "Max Conc",
        CASE s.matrix_code
            WHEN 'SO' THEN sl.rsl_residential_soil_mg_kg
            WHEN 'GW' THEN sl.rsl_residential_tap_ug_l
        END as "RSL",
        ROUND(MAX(r.result_value) / NULLIF(
            CASE s.matrix_code
                WHEN 'SO' THEN sl.rsl_residential_soil_mg_kg
                WHEN 'GW' THEN sl.rsl_residential_tap_ug_l
            END, 0), 2) as "Max HQ",
        sl.carcinogen as "Carcinogen",
        sl.target_organ as "Target Organ"
    FROM fact_results r
    JOIN fact_samples s ON r.sample_id = s.sample_id
    LEFT JOIN dim_analytes a ON r.cas_rn = a.cas_rn
    LEFT JOIN dim_matrix m ON s.matrix_code = m.matrix_code
    LEFT JOIN ref_screening_levels sl ON r.cas_rn = sl.cas_rn
    WHERE r.detect_flag = 'Y'
    AND r.result_value > CASE s.matrix_code
        WHEN 'SO' THEN sl.rsl_residential_soil_mg_kg
        WHEN 'GW' THEN sl.rsl_residential_tap_ug_l
    END
    GROUP BY a.analyte_name, a.analyte_group, m.matrix_name, s.matrix_code,
             sl.rsl_residential_soil_mg_kg, sl.rsl_residential_tap_ug_l, sl.carcinogen, sl.target_organ
    ORDER BY MAX(r.result_value) / NULLIF(CASE s.matrix_code
        WHEN 'SO' THEN sl.rsl_residential_soil_mg_kg
        WHEN 'GW' THEN sl.rsl_residential_tap_ug_l
    END, 0) DESC
    """
    copc_df = conn.execute(copc_query).fetchdf()

    for r_idx, row in enumerate(dataframe_to_rows(copc_df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border

    style_header(ws4, start_row)
    auto_width(ws4)

    # ========================================
    # Table 5: EPCs (calculated inline)
    # ========================================
    ws5 = wb.create_sheet("Table 5 - EPCs")
    start_row = add_title(ws5, "Table 5: Exposure Point Concentrations", site_name.value, report_date.value)

    # Get raw data for EPC calculation
    epc_raw_query = """
    SELECT
        a.analyte_name,
        m.matrix_name,
        r.cas_rn,
        s.matrix_code,
        r.result_value,
        r.detection_limit,
        r.detect_flag,
        r.result_unit,
        CASE s.matrix_code
            WHEN 'SO' THEN sl.rsl_residential_soil_mg_kg
            WHEN 'GW' THEN sl.rsl_residential_tap_ug_l
        END as screening_level
    FROM fact_results r
    JOIN fact_samples s ON r.sample_id = s.sample_id
    LEFT JOIN dim_analytes a ON r.cas_rn = a.cas_rn
    LEFT JOIN dim_matrix m ON s.matrix_code = m.matrix_code
    LEFT JOIN ref_screening_levels sl ON r.cas_rn = sl.cas_rn
    WHERE r.detect_flag = 'Y'  -- Only detects for EPC
    ORDER BY a.analyte_name, s.matrix_code
    """
    epc_raw = conn.execute(epc_raw_query).fetchdf()

    # Calculate EPCs
    epc_results = []
    for (cas, matrix), group in epc_raw.groupby(['cas_rn', 'matrix_code']):
        analyte = group['analyte_name'].iloc[0]
        matrix_nm = group['matrix_name'].iloc[0]
        unit = group['result_unit'].iloc[0]
        sl_val = group['screening_level'].iloc[0]

        vals = group['result_value'].values
        n = len(vals)

        if n >= 4:
            mean_v = np.mean(vals)
            std_v = np.std(vals, ddof=1)
            se = std_v / np.sqrt(n)
            t_crit = stats.t.ppf(0.95, n - 1)
            epc = mean_v + t_crit * se
            method = "95% UCL (t)"
        elif n > 0:
            epc = np.max(vals)
            method = "Maximum"
        else:
            continue

        hq = epc / sl_val if sl_val else None

        epc_results.append({
            'Analyte': analyte,
            'Matrix': matrix_nm,
            'N': n,
            'Mean': round(np.mean(vals), 4),
            'Max': round(np.max(vals), 4),
            'EPC': round(epc, 4),
            'Unit': unit,
            'Method': method,
            'RSL': sl_val,
            'HQ': round(hq, 2) if hq else None
        })

    epc_df = pd.DataFrame(epc_results)

    if len(epc_df) > 0:
        for r_idx, row in enumerate(dataframe_to_rows(epc_df, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, 1):
                cell = ws5.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                # Highlight HQ > 1
                if c_idx == 10 and isinstance(value, (int, float)) and value > 1:
                    for c in range(1, 11):
                        ws5.cell(row=r_idx, column=c).fill = exceed_fill

        style_header(ws5, start_row)
    auto_width(ws5)

    # ========================================
    # Save workbook
    # ========================================
    output_file = OUTPUT_DIR / f"ERA_Summary_Tables_{today}.xlsx"
    wb.save(output_file)

    mo.md(f"""
    ## Report Generated Successfully!

    **File**: `{output_file.name}`
    **Location**: `data/output/`

    ### Tables Included:
    1. **Sample Summary** - {len(sample_df)} samples
    2. **Detection Summary** - {len(detection_df)} analyte/matrix combinations
    3. **Screening Comparison** - {len(screening_df)} comparisons
    4. **COPC Summary** - {len(copc_df)} COPCs identified
    5. **Exposure Point Concentrations** - {len(epc_df)} EPCs calculated

    Yellow highlighting indicates exceedances of screening levels.
    """)
    return (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
        Workbook,
        add_title,
        analyte,
        auto_width,
        c,
        c_idx,
        cas,
        cell,
        copc_df,
        copc_query,
        dataframe_to_rows,
        detection_df,
        detection_query,
        epc,
        epc_df,
        epc_raw,
        epc_raw_query,
        epc_results,
        exceed_fill,
        group,
        header_fill,
        header_font,
        hq,
        matrix,
        matrix_nm,
        mean_v,
        method,
        n,
        output_file,
        r_idx,
        row,
        sample_df,
        sample_query,
        screening_df,
        screening_query,
        se,
        sl_val,
        start_row,
        std_v,
        style_header,
        t_crit,
        thin_border,
        title_font,
        unit,
        vals,
        value,
        wb,
        ws1,
        ws2,
        ws3,
        ws4,
        ws5,
    )


@app.cell
def __(mo):
    mo.md("## Preview Generated Tables")
    return


@app.cell
def __(OUTPUT_DIR, generate_btn, mo, pd, today):
    mo.stop(not generate_btn.value)

    output_path = OUTPUT_DIR / f"ERA_Summary_Tables_{today}.xlsx"

    if output_path.exists():
        preview_data = pd.read_excel(output_path, sheet_name=None)

        tabs = {}
        for sheet_name, df in preview_data.items():
            # Skip title rows
            if len(df) > 0:
                tabs[sheet_name] = mo.ui.table(df.head(20))

        mo.ui.tabs(tabs)
    return df, output_path, preview_data, sheet_name, tabs


@app.cell
def __(mo):
    mo.md(
        r"""
        ## Workflow Complete!

        You have successfully:

        1. **Ingested** lab EDD data from Excel into DuckDB
        2. **Compared** results to EPA screening levels
        3. **Calculated** statistical EPCs using appropriate methods
        4. **Generated** formal ERA summary tables

        ### Key Outputs

        - `ERA_Summary_Tables_YYYYMMDD.xlsx` - Complete ERA data package

        ### Notes for Use

        - Review COPCs and verify against project-specific criteria
        - Consider background concentrations when evaluating site data
        - Use ProUCL for formal EPC calculations in regulatory submissions
        - Update RSL reference table with current EPA values before each project

        ### Data Management

        All source data remains in DuckDB at:
        `data/processed/analytics.duckdb`

        This can be queried directly for additional analyses.
        """
    )
    return


if __name__ == "__main__":
    app.run()
