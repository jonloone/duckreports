#!/usr/bin/env python3
"""
Create the Monthly Report Excel template.
Run after installing dependencies.

Usage:
    python scripts/create_template.py
"""

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    exit(1)

# Output directory
TEMPLATE_DIR = Path(__file__).parent.parent / "templates"
TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)


def create_monthly_report_template():
    """Create the monthly report Excel template."""
    wb = Workbook()

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Title
    ws_summary["A1"] = "Monthly Report"
    ws_summary["A1"].font = title_font
    ws_summary["A2"] = "Report Date: {{date}}"
    ws_summary["A3"] = "Period: {{period}}"

    # Summary table headers
    summary_headers = ["Metric", "Value", "Change"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=5, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Sample metrics
    metrics = [
        ("Total Orders", "{{total_orders}}", "{{orders_change}}"),
        ("Total Revenue", "{{total_revenue}}", "{{revenue_change}}"),
        ("Unique Customers", "{{unique_customers}}", "{{customers_change}}"),
        ("Avg Order Value", "{{avg_order_value}}", "{{aov_change}}"),
    ]

    for row, (metric, value, change) in enumerate(metrics, 6):
        ws_summary.cell(row=row, column=1, value=metric).border = thin_border
        ws_summary.cell(row=row, column=2, value=value).border = thin_border
        ws_summary.cell(row=row, column=3, value=change).border = thin_border

    # Column widths
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 15
    ws_summary.column_dimensions["C"].width = 12

    # Sheet 2: Details
    ws_details = wb.create_sheet(title="Details")

    detail_headers = ["Order ID", "Customer", "Product", "Quantity", "Amount", "Date"]
    for col, header in enumerate(detail_headers, 1):
        cell = ws_details.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Placeholder rows
    for row in range(2, 7):
        for col in range(1, 7):
            cell = ws_details.cell(row=row, column=col, value="{{data}}")
            cell.border = thin_border

    # Column widths
    for col, width in [("A", 10), ("B", 20), ("C", 20), ("D", 10), ("E", 12), ("F", 12)]:
        ws_details.column_dimensions[col].width = width

    # Sheet 3: Charts (placeholder)
    ws_charts = wb.create_sheet(title="Charts")
    ws_charts["A1"] = "Charts will be generated here"
    ws_charts["A3"] = "Use openpyxl's chart module to create:"
    ws_charts["A4"] = "- Revenue trend line chart"
    ws_charts["A5"] = "- Orders by category pie chart"
    ws_charts["A6"] = "- Top customers bar chart"

    # Save template
    output_path = TEMPLATE_DIR / "Monthly_Report_Template.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")


def main():
    print("Creating Excel templates...")
    print("-" * 40)
    create_monthly_report_template()
    print("-" * 40)
    print("Template creation complete!")


if __name__ == "__main__":
    main()
