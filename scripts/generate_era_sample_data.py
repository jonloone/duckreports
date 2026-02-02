#!/usr/bin/env python3
"""
Generate realistic sample ERA data in EDD-like Excel format.
This simulates lab data deliverables from a contaminated site assessment.

Usage:
    python scripts/generate_era_sample_data.py
"""

import random
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    exit(1)

# Set random seed for reproducibility
random.seed(42)

PROJECT_ROOT = Path(__file__).parent.parent
DATA_DIR = PROJECT_ROOT / "data" / "raw"
DATA_DIR.mkdir(parents=True, exist_ok=True)

# Site configuration - simulating a former industrial site
SITE_NAME = "Former Manufacturing Facility"

# Monitoring locations
LOCATIONS = [
    # Monitoring Wells (groundwater)
    {"id": "MW-01", "name": "Monitoring Well 01", "type": "MW", "lat": 40.7128, "lon": -74.0060, "depth": 35},
    {"id": "MW-02", "name": "Monitoring Well 02", "type": "MW", "lat": 40.7130, "lon": -74.0058, "depth": 40},
    {"id": "MW-03", "name": "Monitoring Well 03", "type": "MW", "lat": 40.7125, "lon": -74.0062, "depth": 32},
    {"id": "MW-04", "name": "Monitoring Well 04", "type": "MW", "lat": 40.7132, "lon": -74.0055, "depth": 38},
    {"id": "MW-05", "name": "Monitoring Well 05 (Background)", "type": "MW", "lat": 40.7140, "lon": -74.0070, "depth": 30},
    # Soil Borings
    {"id": "SB-01", "name": "Soil Boring 01", "type": "SB", "lat": 40.7127, "lon": -74.0059, "depth": 20},
    {"id": "SB-02", "name": "Soil Boring 02", "type": "SB", "lat": 40.7129, "lon": -74.0057, "depth": 15},
    {"id": "SB-03", "name": "Soil Boring 03", "type": "SB", "lat": 40.7126, "lon": -74.0061, "depth": 18},
    {"id": "SB-04", "name": "Soil Boring 04", "type": "SB", "lat": 40.7131, "lon": -74.0056, "depth": 22},
    {"id": "SB-05", "name": "Soil Boring 05 (Background)", "type": "SB", "lat": 40.7142, "lon": -74.0068, "depth": 12},
]

# Analytes to test (CAS, name, detection_limit, unit, typical_range, contamination_factor)
# contamination_factor: higher = more likely to have elevated results at contaminated locations
SOIL_ANALYTES = [
    # Metals
    ("7440-38-2", "Arsenic", 1.0, "mg/kg", (2, 15), 1.5),
    ("7440-39-3", "Barium", 5.0, "mg/kg", (50, 300), 1.2),
    ("7440-43-9", "Cadmium", 0.5, "mg/kg", (0.1, 2), 2.0),
    ("7439-92-1", "Lead", 5.0, "mg/kg", (10, 100), 3.0),
    ("7439-97-6", "Mercury", 0.05, "mg/kg", (0.02, 0.1), 1.5),
    ("7440-02-0", "Nickel", 2.0, "mg/kg", (10, 50), 1.3),
    ("7440-66-6", "Zinc", 10.0, "mg/kg", (30, 200), 1.5),
    # VOCs
    ("71-43-2", "Benzene", 0.005, "mg/kg", (0.001, 0.1), 5.0),
    ("100-41-4", "Ethylbenzene", 0.005, "mg/kg", (0.001, 0.05), 3.0),
    ("108-88-3", "Toluene", 0.005, "mg/kg", (0.001, 0.2), 4.0),
    ("1330-20-7", "Xylenes (Total)", 0.005, "mg/kg", (0.001, 0.3), 4.0),
    ("127-18-4", "Tetrachloroethylene (PCE)", 0.005, "mg/kg", (0.001, 0.5), 8.0),
    ("79-01-6", "Trichloroethylene (TCE)", 0.005, "mg/kg", (0.001, 0.3), 8.0),
    # SVOCs
    ("91-20-3", "Naphthalene", 0.33, "mg/kg", (0.1, 2), 3.0),
    ("50-32-8", "Benzo(a)pyrene", 0.033, "mg/kg", (0.01, 0.5), 2.5),
    # PCBs
    ("PCB-TOTAL", "PCBs (Total)", 0.1, "mg/kg", (0.05, 1), 4.0),
]

GROUNDWATER_ANALYTES = [
    # Metals
    ("7440-38-2", "Arsenic", 1.0, "ug/L", (1, 10), 1.5),
    ("7440-39-3", "Barium", 10.0, "ug/L", (20, 200), 1.2),
    ("7439-92-1", "Lead", 1.0, "ug/L", (1, 15), 2.0),
    # VOCs
    ("71-43-2", "Benzene", 0.5, "ug/L", (0.1, 5), 5.0),
    ("100-41-4", "Ethylbenzene", 0.5, "ug/L", (0.1, 3), 3.0),
    ("108-88-3", "Toluene", 0.5, "ug/L", (0.1, 5), 4.0),
    ("1330-20-7", "Xylenes (Total)", 0.5, "ug/L", (0.1, 8), 4.0),
    ("127-18-4", "Tetrachloroethylene (PCE)", 0.5, "ug/L", (0.1, 50), 10.0),
    ("79-01-6", "Trichloroethylene (TCE)", 0.5, "ug/L", (0.1, 30), 10.0),
    ("75-01-4", "Vinyl Chloride", 0.2, "ug/L", (0.05, 5), 8.0),
    # PFAS
    ("335-67-1", "PFOA", 0.002, "ug/L", (0.001, 0.05), 5.0),
    ("1763-23-1", "PFOS", 0.002, "ug/L", (0.001, 0.05), 5.0),
]

# Sampling events (quarters)
SAMPLE_EVENTS = [
    datetime(2024, 3, 15),   # Q1 2024
    datetime(2024, 6, 20),   # Q2 2024
    datetime(2024, 9, 18),   # Q3 2024
    datetime(2024, 12, 10),  # Q4 2024
]


def generate_result(analyte, is_contaminated_location, is_groundwater=False):
    """Generate a realistic analytical result."""
    cas, name, dl, unit, typical_range, contam_factor = analyte

    # Determine if this will be a detect or non-detect
    # Background locations have lower detection probability
    base_detect_prob = 0.6 if is_contaminated_location else 0.3

    # Metals almost always detected, VOCs less so
    if cas.startswith("744") or cas.startswith("743"):  # Metals
        detect_prob = 0.95
    elif "PCB" in cas:
        detect_prob = 0.3 if is_contaminated_location else 0.05
    elif cas.startswith("335") or cas.startswith("1763"):  # PFAS
        detect_prob = 0.4 if is_contaminated_location else 0.1
    else:  # VOCs, SVOCs
        detect_prob = base_detect_prob

    if random.random() > detect_prob:
        # Non-detect
        return dl, "N", "U"

    # Generate a detected value
    low, high = typical_range

    if is_contaminated_location:
        # Higher values at contaminated locations
        value = random.uniform(low, high * contam_factor)
        # Occasionally generate values above screening levels
        if random.random() < 0.2:
            value *= random.uniform(2, 10)
    else:
        # Background - typically lower
        value = random.uniform(low * 0.5, high * 0.5)

    # Round to appropriate precision
    if value < 0.01:
        value = round(value, 4)
    elif value < 1:
        value = round(value, 3)
    elif value < 100:
        value = round(value, 2)
    else:
        value = round(value, 1)

    # Determine qualifier
    qualifier = ""
    if value < dl * 2 and value >= dl:
        # J flag - estimated (between DL and QL)
        qualifier = "J"
    elif random.random() < 0.02:
        # Occasional other qualifiers
        qualifier = random.choice(["B", "E", "D"])

    return value, "Y", qualifier


def create_locations_excel():
    """Create locations Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Locations"

    headers = ["location_id", "location_name", "location_type", "latitude", "longitude",
               "elevation_ft", "total_depth_ft", "install_date", "status"]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row, loc in enumerate(LOCATIONS, 2):
        ws.cell(row=row, column=1, value=loc["id"])
        ws.cell(row=row, column=2, value=loc["name"])
        ws.cell(row=row, column=3, value=loc["type"])
        ws.cell(row=row, column=4, value=loc["lat"])
        ws.cell(row=row, column=5, value=loc["lon"])
        ws.cell(row=row, column=6, value=random.uniform(95, 105))  # elevation
        ws.cell(row=row, column=7, value=loc["depth"])
        ws.cell(row=row, column=8, value="2023-06-15")
        ws.cell(row=row, column=9, value="Active")

    # Adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    output_path = DATA_DIR / "site_locations.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")


def create_lab_results_excel():
    """Create lab results Excel file (EDD-style format)."""
    wb = Workbook()

    # Sheet 1: Samples
    ws_samples = wb.active
    ws_samples.title = "Samples"

    sample_headers = ["sample_id", "location_id", "sample_date", "sample_time", "matrix_code",
                      "sample_type", "depth_top_ft", "depth_bottom_ft", "sample_method",
                      "sampler_name", "lab_name", "lab_sample_id"]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(sample_headers, 1):
        cell = ws_samples.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Sheet 2: Results
    ws_results = wb.create_sheet(title="Results")

    result_headers = ["sample_id", "cas_rn", "analyte_name", "result_value", "result_unit",
                      "detection_limit", "detect_flag", "lab_qualifier", "dilution_factor",
                      "analysis_method", "analysis_date", "basis", "percent_moisture"]

    for col, header in enumerate(result_headers, 1):
        cell = ws_results.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    sample_row = 2
    result_row = 2
    sample_counter = 1

    for event_date in SAMPLE_EVENTS:
        for loc in LOCATIONS:
            # Determine if this is a contaminated location (not background)
            is_contaminated = "Background" not in loc["name"]

            # Determine matrix and analytes based on location type
            if loc["type"] == "MW":
                matrix = "GW"
                analytes = GROUNDWATER_ANALYTES
                depth_top = loc["depth"] - 10
                depth_bottom = loc["depth"]
                method = "Low-Flow"
            else:
                matrix = "SO"
                analytes = SOIL_ANALYTES
                # Multiple depths for soil borings
                depths = [(0, 2), (4, 6), (8, 10)] if is_contaminated else [(0, 2), (4, 6)]

            # For soil, create samples at multiple depths
            sample_depths = [(None, None)] if matrix == "GW" else depths

            for depth_top, depth_bottom in sample_depths:
                sample_id = f"{loc['id']}-{event_date.strftime('%Y%m%d')}"
                if depth_top is not None:
                    sample_id += f"-{int(depth_top)}-{int(depth_bottom)}"

                lab_sample_id = f"LAB{sample_counter:05d}"

                # Write sample record
                ws_samples.cell(row=sample_row, column=1, value=sample_id)
                ws_samples.cell(row=sample_row, column=2, value=loc["id"])
                ws_samples.cell(row=sample_row, column=3, value=event_date.strftime("%Y-%m-%d"))
                ws_samples.cell(row=sample_row, column=4, value="10:30")
                ws_samples.cell(row=sample_row, column=5, value=matrix)
                ws_samples.cell(row=sample_row, column=6, value="N")  # Normal sample
                ws_samples.cell(row=sample_row, column=7, value=depth_top)
                ws_samples.cell(row=sample_row, column=8, value=depth_bottom)
                ws_samples.cell(row=sample_row, column=9, value="Low-Flow" if matrix == "GW" else "Direct Push")
                ws_samples.cell(row=sample_row, column=10, value="J. Smith")
                ws_samples.cell(row=sample_row, column=11, value="TestAmerica")
                ws_samples.cell(row=sample_row, column=12, value=lab_sample_id)

                sample_row += 1
                sample_counter += 1

                # Generate results for each analyte
                for analyte in analytes:
                    cas, name, dl, unit, typical_range, contam_factor = analyte
                    result_value, detect_flag, qualifier = generate_result(
                        analyte, is_contaminated, matrix == "GW"
                    )

                    # Determine analysis method
                    if cas.startswith("744") or cas.startswith("743"):
                        method = "SW6010D" if matrix == "SO" else "SW6020B"
                    elif cas in ["71-43-2", "100-41-4", "108-88-3", "1330-20-7", "127-18-4", "79-01-6", "75-01-4"]:
                        method = "SW8260D"
                    elif cas in ["91-20-3", "50-32-8"]:
                        method = "SW8270E"
                    elif "PCB" in cas:
                        method = "SW8082A"
                    elif cas in ["335-67-1", "1763-23-1"]:
                        method = "SW537.1"
                    else:
                        method = "SW8260D"

                    analysis_date = event_date + timedelta(days=random.randint(5, 14))

                    # Write result record
                    ws_results.cell(row=result_row, column=1, value=sample_id)
                    ws_results.cell(row=result_row, column=2, value=cas)
                    ws_results.cell(row=result_row, column=3, value=name)
                    ws_results.cell(row=result_row, column=4, value=result_value)
                    ws_results.cell(row=result_row, column=5, value=unit)
                    ws_results.cell(row=result_row, column=6, value=dl)
                    ws_results.cell(row=result_row, column=7, value=detect_flag)
                    ws_results.cell(row=result_row, column=8, value=qualifier)
                    ws_results.cell(row=result_row, column=9, value=1)  # dilution
                    ws_results.cell(row=result_row, column=10, value=method)
                    ws_results.cell(row=result_row, column=11, value=analysis_date.strftime("%Y-%m-%d"))
                    ws_results.cell(row=result_row, column=12, value="Dry" if matrix == "SO" else "")
                    ws_results.cell(row=result_row, column=13,
                                    value=round(random.uniform(5, 25), 1) if matrix == "SO" else "")

                    result_row += 1

    # Adjust column widths
    for ws in [ws_samples, ws_results]:
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

    output_path = DATA_DIR / "lab_results_edd.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")
    print(f"  Samples: {sample_row - 2}")
    print(f"  Results: {result_row - 2}")


def create_field_measurements_excel():
    """Create field measurements Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Field_Measurements"

    headers = ["sample_id", "location_id", "measurement_date", "parameter", "result", "unit", "notes"]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    row = 2
    for event_date in SAMPLE_EVENTS:
        for loc in LOCATIONS:
            if loc["type"] == "MW":  # Only for groundwater
                sample_id = f"{loc['id']}-{event_date.strftime('%Y%m%d')}"

                # Field parameters
                params = [
                    ("pH", round(random.uniform(6.0, 8.0), 2), "SU"),
                    ("Specific Conductance", round(random.uniform(200, 800), 0), "umhos/cm"),
                    ("Temperature", round(random.uniform(12, 18), 1), "deg C"),
                    ("Dissolved Oxygen", round(random.uniform(2, 8), 2), "mg/L"),
                    ("ORP", round(random.uniform(-100, 200), 0), "mV"),
                    ("Turbidity", round(random.uniform(1, 50), 1), "NTU"),
                    ("Depth to Water", round(random.uniform(5, 15), 2), "ft"),
                ]

                for param, value, unit in params:
                    ws.cell(row=row, column=1, value=sample_id)
                    ws.cell(row=row, column=2, value=loc["id"])
                    ws.cell(row=row, column=3, value=event_date.strftime("%Y-%m-%d"))
                    ws.cell(row=row, column=4, value=param)
                    ws.cell(row=row, column=5, value=value)
                    ws.cell(row=row, column=6, value=unit)
                    ws.cell(row=row, column=7, value="")
                    row += 1

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    output_path = DATA_DIR / "field_measurements.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")


def main():
    print("Generating ERA sample data...")
    print(f"Site: {SITE_NAME}")
    print("-" * 50)

    create_locations_excel()
    create_lab_results_excel()
    create_field_measurements_excel()

    print("-" * 50)
    print("ERA sample data generation complete!")
    print(f"\nFiles created in: {DATA_DIR}")
    print("\nThese files simulate lab EDD data from a contaminated site:")
    print("  - site_locations.xlsx: Monitoring wells and soil borings")
    print("  - lab_results_edd.xlsx: Lab analytical results (Samples + Results sheets)")
    print("  - field_measurements.xlsx: Field parameter measurements")


if __name__ == "__main__":
    main()
