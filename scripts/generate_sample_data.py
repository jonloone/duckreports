#!/usr/bin/env python3
"""
Generate sample Excel files for the Marimo + DuckDB starter.
Run this script after installing dependencies to create sample data.

Usage:
    python scripts/generate_sample_data.py
"""

import random
from datetime import datetime, timedelta
from pathlib import Path

# Ensure openpyxl is available
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    exit(1)

# Set random seed for reproducible sample data
random.seed(42)

# Output directory
DATA_DIR = Path(__file__).parent.parent / "data" / "raw"
DATA_DIR.mkdir(parents=True, exist_ok=True)


def create_customers_excel():
    """Create sample customers Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    # Headers
    headers = ["customer_id", "customer_name", "region", "segment", "email", "created_date"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Sample data
    regions = ["North", "South", "East", "West"]
    segments = ["Enterprise", "SMB", "Consumer"]

    customers = [
        (1001, "Acme Corporation", "North", "Enterprise", "contact@acme.com", "2023-01-15"),
        (1002, "Widget Inc", "South", "SMB", "info@widget.com", "2023-02-20"),
        (1003, "TechStart LLC", "East", "SMB", "hello@techstart.io", "2023-03-10"),
        (1004, "Global Dynamics", "West", "Enterprise", "sales@globaldyn.com", "2023-03-22"),
        (1005, "Local Shop", "North", "Consumer", "owner@localshop.com", "2023-04-05"),
        (1006, "DataFlow Systems", "East", "Enterprise", "contact@dataflow.net", "2023-04-18"),
        (1007, "Creative Agency", "South", "SMB", "team@creative.co", "2023-05-02"),
        (1008, "Metro Services", "West", "Consumer", "support@metro.com", "2023-05-15"),
        (1009, "Alpine Industries", "North", "Enterprise", "info@alpine.com", "2023-06-01"),
        (1010, "Sunshine Retail", "South", "Consumer", "contact@sunshine.com", "2023-06-20"),
    ]

    for row, customer in enumerate(customers, 2):
        for col, value in enumerate(customer, 1):
            ws.cell(row=row, column=col, value=value)

    # Adjust column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 14

    output_path = DATA_DIR / "sample_customers.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")
    return customers


def create_products_excel():
    """Create sample products Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    # Headers
    headers = ["product_id", "product_name", "category", "unit_price", "stock_quantity"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Sample data
    products = [
        (2001, "Widget Pro", "Hardware", 49.99, 150),
        (2002, "Widget Basic", "Hardware", 29.99, 300),
        (2003, "DataSync License", "Software", 199.00, 999),
        (2004, "Support Package", "Services", 99.00, 999),
        (2005, "Premium Cable", "Accessories", 19.99, 500),
        (2006, "Adapter Kit", "Accessories", 14.99, 250),
        (2007, "Analytics Suite", "Software", 299.00, 999),
        (2008, "Widget Enterprise", "Hardware", 149.99, 75),
        (2009, "Training Course", "Services", 249.00, 999),
        (2010, "Mounting Bracket", "Accessories", 9.99, 400),
        (2011, "Security Module", "Software", 149.00, 999),
        (2012, "Extended Warranty", "Services", 49.00, 999),
        (2013, "Widget Compact", "Hardware", 39.99, 200),
        (2014, "Integration Pack", "Software", 79.00, 999),
        (2015, "Carrying Case", "Accessories", 24.99, 350),
    ]

    for row, product in enumerate(products, 2):
        for col, value in enumerate(product, 1):
            ws.cell(row=row, column=col, value=value)

    # Adjust column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14

    output_path = DATA_DIR / "sample_products.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")
    return products


def create_orders_excel(customers, products):
    """Create sample orders Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    # Headers
    headers = ["order_id", "customer_id", "product_id", "quantity", "unit_price", "total_amount", "order_date", "status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Generate 50 orders
    customer_ids = [c[0] for c in customers]
    product_data = {p[0]: p[3] for p in products}  # product_id -> unit_price
    product_ids = list(product_data.keys())
    statuses = ["Completed", "Completed", "Completed", "Pending", "Shipped"]  # Weighted towards completed

    start_date = datetime(2024, 1, 1)
    orders = []

    for i in range(50):
        order_id = 3000 + i + 1
        customer_id = random.choice(customer_ids)
        product_id = random.choice(product_ids)
        quantity = random.randint(1, 10)
        unit_price = product_data[product_id]
        total_amount = round(quantity * unit_price, 2)
        order_date = (start_date + timedelta(days=random.randint(0, 365))).strftime("%Y-%m-%d")
        status = random.choice(statuses)

        orders.append((order_id, customer_id, product_id, quantity, unit_price, total_amount, order_date, status))

    # Sort by order_id
    orders.sort(key=lambda x: x[0])

    for row, order in enumerate(orders, 2):
        for col, value in enumerate(order, 1):
            ws.cell(row=row, column=col, value=value)

    # Adjust column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 10

    output_path = DATA_DIR / "sample_orders.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")


def main():
    """Generate all sample data files."""
    print("Generating sample data files...")
    print("-" * 40)

    customers = create_customers_excel()
    products = create_products_excel()
    create_orders_excel(customers, products)

    print("-" * 40)
    print("Sample data generation complete!")
    print(f"\nFiles created in: {DATA_DIR}")


if __name__ == "__main__":
    main()
